from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font as OpenPyxlFont
from PySide6.QtWidgets import QMessageBox

from interface_grafica.config import CNPJ_ROOT
from transformacao.resumo_global import gerar_resumo_global_dataframe
from utilitarios.text import display_cell, is_year_column_name


_DATE_ONLY_EXPORT_COLUMNS = {"dt_doc", "dt_e_s", "dt_inicio", "dt_fim"}
_RESUMO_GLOBAL_PERIODO_COLUMNS = {
    "ICMS_entr_desacob_periodo",
    "ICMS_saidas_desac_periodo",
    "ICMS_estoque_desac_periodo",
    "Total_periodo",
}
_RESUMO_GLOBAL_TOTAL_LABELS = (
    ("ICMS_entr_desacob", "ICMS entradas"),
    ("ICMS_saidas_desac", "ICMS saidas"),
    ("ICMS_estoque_desac", "ICMS estoque"),
    ("Total", "Total"),
    ("ICMS_entr_desacob_periodo", "ICMS entradas periodo"),
    ("ICMS_saidas_desac_periodo", "ICMS saidas periodo"),
    ("ICMS_estoque_desac_periodo", "ICMS estoque periodo"),
    ("Total_periodo", "Total periodo"),
)


def _valor_data_excel(coluna: str, valor):
    if str(coluna).strip().lower() not in _DATE_ONLY_EXPORT_COLUMNS:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        texto = valor.strip()
        if not texto:
            return None
        for formato in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(texto, formato).date()
            except ValueError:
                continue
    return None


class RelatoriosResumoControllerMixin:
    def _texto_totais_resumo_global(self, df: pl.DataFrame) -> str:
        if df is None or df.is_empty():
            return "Totais ICMS: sem dados no recorte atual."

        partes: list[str] = []
        for coluna, rotulo in _RESUMO_GLOBAL_TOTAL_LABELS:
            if coluna not in df.columns:
                continue
            valor = df.select(
                pl.col(coluna).cast(pl.Float64, strict=False).fill_null(0).sum()
            ).item()
            partes.append(f"{rotulo}: {display_cell(float(valor or 0), coluna)}")

        if not partes:
            return "Totais ICMS: colunas de ICMS nao disponiveis no recorte atual."
        return "Totais ICMS | " + " | ".join(partes)

    def _atualizar_totais_resumo_global(self, df: pl.DataFrame) -> None:
        label = getattr(self, "lbl_resumo_global_totais", None)
        if label is not None and hasattr(label, "setText"):
            label.setText(self._texto_totais_resumo_global(df))

    def _escrever_planilha_openpyxl(self, ws, df: pl.DataFrame) -> None:
        fonte_padrao = OpenPyxlFont(name="Arial", size=8)
        fonte_header = OpenPyxlFont(name="Arial", size=8, bold=True)
        formato_num = "#,##0.00"
        formato_inteiro = "#,##0"
        formato_ano = "0"
        formato_data = "dd/mm/yyyy"
        formato_data_hora = "dd/mm/yyyy hh:mm:ss"

        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.font = fonte_header

        for row in df.iter_rows():
            linha_excel = []
            for coluna, valor in zip(df.columns, row, strict=False):
                data_excel = _valor_data_excel(coluna, valor)
                if data_excel is not None:
                    linha_excel.append(data_excel)
                elif isinstance(valor, (int, float)) and not isinstance(valor, bool):
                    linha_excel.append(valor)
                else:
                    linha_excel.append(display_cell(valor, coluna))
            ws.append(linha_excel)

        for row in ws.iter_rows():
            for cell in row:
                if cell.row == 1:
                    cell.font = fonte_header
                else:
                    cell.font = fonte_padrao
                if cell.row == 1:
                    continue
                nome_coluna = str(ws.cell(row=1, column=cell.column).value or "")
                if (
                    is_year_column_name(nome_coluna)
                    and isinstance(cell.value, (int, float))
                    and not isinstance(cell.value, bool)
                ):
                    cell.number_format = formato_ano
                elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    if float(cell.value).is_integer():
                        cell.number_format = formato_inteiro
                    else:
                        cell.number_format = formato_num
                elif hasattr(cell.value, "hour"):
                    cell.number_format = formato_data_hora
                elif hasattr(cell.value, "day") and hasattr(cell.value, "month"):
                    cell.number_format = formato_data

        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

    def _gerar_resumo_global(
        self,
        mensal: pl.DataFrame,
        anual: pl.DataFrame,
        periodos: pl.DataFrame | list[int] | None = None,
        anos_base: list[int] | None = None,
    ) -> pl.DataFrame:
        if isinstance(periodos, list) and anos_base is None:
            anos_base = periodos
            periodos = None
        return gerar_resumo_global_dataframe(
            mensal,
            anual,
            periodos if isinstance(periodos, pl.DataFrame) else pl.DataFrame(),
            anos_base,
            manter_competencias_zeradas=True,
        )

    def _anos_base_resumo(self, dt_ini, dt_fim) -> list[int] | None:
        try:
            ano_ini_v = int(self.cmb_resumo_global_ano_ini.currentText())
            ano_fim_v = int(self.cmb_resumo_global_ano_fim.currentText())
        except (AttributeError, ValueError):
            intervalo = getattr(self, "_intervalo_anos_produtos_selecionados", None)
            if callable(intervalo):
                ano_ini_v, ano_fim_v = intervalo()
            else:
                ano_ini_v, ano_fim_v = None, None
            ano_ini_v = 2021 if ano_ini_v is None else ano_ini_v
            ano_fim_v = 2025 if ano_fim_v is None else ano_fim_v
        if ano_ini_v > ano_fim_v:
            ano_ini_v, ano_fim_v = ano_fim_v, ano_ini_v
        if dt_ini is not None and dt_fim is not None:
            return list(range(dt_ini.year(), dt_fim.year() + 1))
        if dt_ini is not None:
            return list(range(dt_ini.year(), ano_fim_v + 1))
        if dt_fim is not None:
            return list(range(ano_ini_v, dt_fim.year() + 1))
        return list(range(ano_ini_v, ano_fim_v + 1))

    def _resumo_global_filtrar_selecionados(self) -> bool:
        checkbox = getattr(self, "chk_resumo_global_so_selecionados", None)
        return bool(checkbox is not None and checkbox.isChecked())

    def _filtrar_df_por_ids_resumo(self, df: pl.DataFrame, ids: list[str] | None) -> pl.DataFrame:
        if ids is None or df.is_empty():
            return df
        col = next((c for c in ("id_agregado", "id_agrupado") if c in df.columns), None)
        if col is None:
            return df
        return df.filter(pl.col(col).cast(pl.Utf8, strict=False).is_in(ids))

    def _resumo_global_qdate_ativo(self, attr_name: str):
        widget = getattr(self, attr_name, None)
        if widget is None or not hasattr(widget, "date"):
            return None
        value = widget.date()
        converter = getattr(self, "_valor_qdate_ativo", None)
        if callable(converter):
            return converter(value)
        minimum = widget.minimumDate() if hasattr(widget, "minimumDate") else None
        if minimum is not None and value == minimum:
            return None
        return value

    def atualizar_aba_resumo_global(self) -> None:
        cnpj = self.state.current_cnpj
        if not cnpj:
            return

        dt_ini = self._resumo_global_qdate_ativo("resumo_global_filter_data_ini")
        dt_fim = self._resumo_global_qdate_ativo("resumo_global_filter_data_fim")
        tem_filtro_data = dt_ini is not None or dt_fim is not None

        filtrar_selecionados = self._resumo_global_filtrar_selecionados()
        ids_marcados = (
            self._ids_produtos_selecionados_para_exportacao() if filtrar_selecionados else []
        )
        tem_aba_produtos = (
            filtrar_selecionados
            and hasattr(self, "_produtos_selecionados_df")
            and not self._produtos_selecionados_df.is_empty()
        )

        if tem_aba_produtos and not ids_marcados:
            self.lbl_resumo_global_status.setText(
                "Marque produtos na aba 'Produtos selecionados' para exibir o resumo."
            )
            self._resumo_global_df = pl.DataFrame()
            self.resumo_global_model.set_dataframe(pl.DataFrame())
            self._atualizar_totais_resumo_global(pl.DataFrame())
            return

        if ids_marcados:
            try:
                n = len(ids_marcados)
                ini_str = dt_ini.toString("dd/MM/yyyy") if dt_ini else "inicio"
                fim_str = dt_fim.toString("dd/MM/yyyy") if dt_fim else "fim"
                partes = [f"{n} produtos marcados"]
                if tem_filtro_data:
                    partes.append(f"periodo {ini_str} a {fim_str}")
                self.lbl_resumo_global_status.setText(
                    f"⏳ Consolidando resumo: {', '.join(partes)}..."
                )

                mensal = self._filtrar_dataframe_por_ids(
                    self._produtos_selecionados_mensal_df, ids_marcados
                )
                anual = self._filtrar_dataframe_por_ids(
                    self._produtos_selecionados_anual_df, ids_marcados
                )
                periodos = self._filtrar_dataframe_por_ids(
                    self._produtos_selecionados_periodos_df, ids_marcados
                )

                if tem_filtro_data:
                    mensal = self._filtrar_dataframe_produtos_selecionados_por_data(
                        mensal, dt_ini, dt_fim, "mensal"
                    )
                    anual = self._filtrar_dataframe_produtos_selecionados_por_data(
                        anual, dt_ini, dt_fim, "anual"
                    )

                anos_base = self._anos_base_resumo(dt_ini, dt_fim)
                resumo = self._gerar_resumo_global(mensal, anual, periodos, anos_base)
                self._resumo_global_df = resumo
                self.resumo_global_model.set_dataframe(resumo)
                self._atualizar_totais_resumo_global(resumo)

                partes_label = [f"Resumo global: {n} produtos marcados"]
                if tem_filtro_data:
                    partes_label.append(f"periodo {ini_str} a {fim_str}")
                partes_label.append(f"{resumo.height} competencias")
                self.lbl_resumo_global_status.setText(", ".join(partes_label) + ".")
                return
            except Exception as e:
                self.show_error("Erro de consolidacao", f"Erro ao filtrar Resumo Global: {e}")
                return

        # Sem produtos selecionados: comportamento anterior (parquet ou calculo em tempo real)
        path = CNPJ_ROOT / cnpj / "analises" / "produtos" / f"aba_resumo_global_{cnpj}.parquet"

        def _finalizar_carga_resumo(df: pl.DataFrame | None, uniques: dict | None = None) -> None:
            if df is None:
                self._resumo_global_df = pl.DataFrame()
                self.resumo_global_model.set_dataframe(pl.DataFrame())
                self._atualizar_totais_resumo_global(pl.DataFrame())
                self.lbl_resumo_global_status.setText(
                    "Resumo global nao encontrado para este CNPJ."
                )
                return
            if (
                not _RESUMO_GLOBAL_PERIODO_COLUMNS.issubset(set(df.columns))
                and not self._aba_mensal_df.is_empty()
                and not self._aba_anual_df.is_empty()
                and {"ano", "mes", "ICMS_entr_desacob"}.issubset(set(self._aba_mensal_df.columns))
                and {"ano", "ICMS_saidas_desac", "ICMS_estoque_desac"}.issubset(
                    set(self._aba_anual_df.columns)
                )
                and (
                    "ICMS_entr_desacob_periodo" in self._aba_mensal_df.columns
                    or {
                        "ICMS_saidas_desac_periodo",
                        "ICMS_estoque_desac_periodo",
                    }.issubset(set(self._aba_periodos_df.columns))
                )
            ):
                df = self._gerar_resumo_global(
                    self._aba_mensal_df,
                    self._aba_anual_df,
                    self._aba_periodos_df,
                    self._anos_base_resumo(dt_ini, dt_fim),
                )
            df = self._aplicar_filtro_data_resumo_global(df, dt_ini, dt_fim)
            self._resumo_global_df = df
            self.resumo_global_model.set_dataframe(df)
            self._atualizar_totais_resumo_global(df)
            self.lbl_resumo_global_status.setText(
                f"Resumo global carregado (total {df.height} competencias)."
            )

        if path.exists():
            self.lbl_resumo_global_status.setText("⏳ Carregando resumo global...")
            self._carregar_dados_parquet_async(
                path, _finalizar_carga_resumo, "Carregando Resumo Global"
            )
            return

        if self._aba_mensal_df.is_empty() or self._aba_anual_df.is_empty():
            path_mensal = CNPJ_ROOT / cnpj / "analises" / "produtos" / f"aba_mensal_{cnpj}.parquet"
            path_anual = CNPJ_ROOT / cnpj / "analises" / "produtos" / f"aba_anual_{cnpj}.parquet"

            needs_load = False
            if self._aba_mensal_df.is_empty() and path_mensal.exists():
                self.atualizar_aba_mensal()
                needs_load = True
            if self._aba_anual_df.is_empty() and path_anual.exists():
                self.atualizar_aba_anual()
                needs_load = True

            if needs_load:
                self.lbl_resumo_global_status.setText(
                    "⏳ Carregando dependencias (Mensal/Anual)... Ao terminar, o resumo sera atualizado."
                )
                return
            else:
                self.resumo_global_model.set_dataframe(pl.DataFrame())
                self._resumo_global_df = pl.DataFrame()
                self._atualizar_totais_resumo_global(pl.DataFrame())
                self.lbl_resumo_global_status.setText(
                    "Aguarde o processamento das abas Mensal e Anual."
                )
                return

        try:
            mensal = self._aba_mensal_df
            anual = self._aba_anual_df
            periodos = self._aba_periodos_df
            if tem_filtro_data:
                mensal = self._filtrar_dataframe_produtos_selecionados_por_data(
                    mensal, dt_ini, dt_fim, "mensal"
                )
                anual = self._filtrar_dataframe_produtos_selecionados_por_data(
                    anual, dt_ini, dt_fim, "anual"
                )
            anos_base = self._anos_base_resumo(dt_ini, dt_fim)
            resumo = self._gerar_resumo_global(mensal, anual, periodos, anos_base)
            self._resumo_global_df = resumo
            self.resumo_global_model.set_dataframe(resumo)
            self._atualizar_totais_resumo_global(resumo)
            self.lbl_resumo_global_status.setText(
                f"Resumo global consolidado ({resumo.height} competencias)."
            )
        except Exception as e:
            self.show_error("Erro de consolidacao", f"Erro ao calcular Resumo Global: {e}")

    def _aplicar_filtro_data_resumo_global(self, df: pl.DataFrame, dt_ini, dt_fim) -> pl.DataFrame:
        if df.is_empty() or "Ano/Mes" not in df.columns:
            return df
        col = pl.col("Ano/Mes").str.strptime(pl.Date, "%Y-%m", strict=False)
        if dt_ini is not None:
            limite = pl.lit(dt_ini.toPython().replace(day=1))
            df = df.filter(col >= limite)
        if dt_fim is not None:
            limite = pl.lit(dt_fim.toPython().replace(day=1))
            df = df.filter(col <= limite)
        if dt_ini is None and dt_fim is None:
            anos = self._anos_base_resumo(None, None)
            if anos:
                df = df.filter(col.dt.year().is_between(min(anos), max(anos)))
        return df

    def exportar_resumo_global_excel(self) -> None:
        if self._resumo_global_df is None or self._resumo_global_df.is_empty():
            QMessageBox.information(self, "Exportacao", "Nao ha dados globais para exportar.")
            return
        target = self._save_dialog("Exportar Resumo Global", "Excel (*.xlsx)")
        if not target:
            return
        try:
            df_to_export = self._dataframe_colunas_perfil(
                "resumo_global",
                "resumo_global",
                self.resumo_global_model,
                self._resumo_global_df,
                perfil="Exportar",
            )
            wb = Workbook()
            ws = wb.active
            ws.title = "Resumo Global"
            self._escrever_planilha_openpyxl(ws, df_to_export)
            target.parent.mkdir(parents=True, exist_ok=True)
            wb.save(target)
            self.show_info("Exportacao concluida", f"Arquivo gerado em:\n{target}")
        except Exception as e:
            self.show_error("Erro de exportacao", str(e))

    def _montar_valores_consolidados_produtos_selecionados(self, ids: list[str]) -> pl.DataFrame:
        resumo = self._filtrar_dataframe_por_ids(self._produtos_selecionados_df, ids)
        mensal = self._filtrar_dataframe_por_ids(self._produtos_selecionados_mensal_df, ids)
        anual = self._filtrar_dataframe_por_ids(self._produtos_selecionados_anual_df, ids)

        if resumo.is_empty():
            return pl.DataFrame(
                schema={
                    "Ano/Mes": pl.Utf8,
                    "ICMS_entr_desacob": pl.Float64,
                    "ICMS_saidas_desac": pl.Float64,
                    "ICMS_estoque_desac": pl.Float64,
                    "Total": pl.Float64,
                }
            )

        anos_base: list[int] = []
        ano_ini, ano_fim = self._intervalo_anos_produtos_selecionados()
        if ano_ini is not None and ano_fim is not None:
            anos_base = list(range(int(ano_ini), int(ano_fim) + 1))
        else:
            anos_base = None

        periodos = self._filtrar_dataframe_por_ids(self._produtos_selecionados_periodos_df, ids)
        return self._gerar_resumo_global(mensal, anual, periodos, anos_base)

    def exportar_produtos_selecionados_excel(self) -> None:
        if self._produtos_selecionados_df.is_empty():
            QMessageBox.information(self, "Exportacao", "Nao ha dados consolidados para exportar.")
            return

        from interface_grafica.ui.dialogs import DialogoExportacaoEstoque

        dlg = DialogoExportacaoEstoque(self)
        if not dlg.exec():
            return
        dt_ini, dt_fim = dlg.obter_datas()

        target = self._save_dialog("Exportar mov_estoque, mensal e anual", "Excel (*.xlsx)")
        if not target:
            return
        try:
            ids = self._ids_produtos_selecionados_para_exportacao()
            if not ids:
                QMessageBox.information(
                    self,
                    "Exportacao",
                    "Marque pelo menos um id_agregado em Produtos selecionados.",
                )
                return

            df_mensal_export = self._filtrar_dataframe_produtos_selecionados_por_data(
                self._produtos_selecionados_mensal_df, dt_ini, dt_fim, "mensal"
            )
            df_anual_export = self._filtrar_dataframe_produtos_selecionados_por_data(
                self._produtos_selecionados_anual_df, dt_ini, dt_fim, "anual"
            )
            df_mov_export = self._filtrar_dataframe_produtos_selecionados_por_data(
                self._produtos_selecionados_mov_df, dt_ini, dt_fim, "mov_estoque"
            )

            mensal = self._dataframe_colunas_perfil(
                "aba_mensal",
                "aba_mensal",
                self.aba_mensal_model,
                self._filtrar_dataframe_por_ids(df_mensal_export, ids),
                perfil="Exportar",
            )
            anual = self._dataframe_colunas_perfil(
                "aba_anual",
                "aba_anual",
                self.aba_anual_model,
                self._filtrar_dataframe_por_ids(df_anual_export, ids),
                perfil="Exportar",
            )
            mov = self._dataframe_colunas_perfil(
                "mov_estoque",
                "mov_estoque",
                self.mov_estoque_model,
                self._filtrar_dataframe_por_ids(df_mov_export, ids),
                perfil="Exportar",
            )
            periodos = self._dataframe_colunas_perfil(
                "aba_periodos",
                "aba_periodos",
                self.aba_periodos_model,
                self._filtrar_dataframe_por_ids(self._produtos_selecionados_periodos_df, ids),
                perfil="Exportar",
            )
            valores_consolidados = self._montar_valores_consolidados_produtos_selecionados(ids)

            produtos_selecionados_tabela = self._dataframe_colunas_perfil(
                "produtos_selecionados",
                "produtos_selecionados",
                self.produtos_selecionados_model,
                self._filtrar_dataframe_por_ids(self._produtos_selecionados_df, ids),
                perfil="Exportar",
            )

            wb = Workbook()
            ws_produtos = wb.active
            ws_produtos.title = "Produtos_Selecionados"
            self._escrever_planilha_openpyxl(ws_produtos, produtos_selecionados_tabela)

            self._escrever_planilha_openpyxl(wb.create_sheet("Mov_Estoque"), mov)
            self._escrever_planilha_openpyxl(wb.create_sheet("Mensal"), mensal)
            self._escrever_planilha_openpyxl(wb.create_sheet("Anual"), anual)
            self._escrever_planilha_openpyxl(wb.create_sheet("Periodos"), periodos)
            self._escrever_planilha_openpyxl(wb.create_sheet("ICMS_devido"), valores_consolidados)
            target.parent.mkdir(parents=True, exist_ok=True)
            wb.save(target)
            self.show_info("Exportacao concluida", f"Arquivo gerado em:\n{target}")
        except Exception as e:
            self.show_error("Erro de exportacao", str(e))
