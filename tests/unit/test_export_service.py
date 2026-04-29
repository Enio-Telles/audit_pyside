from __future__ import annotations

from pathlib import Path
from datetime import date

import polars as pl
import pytest
from openpyxl import load_workbook

from interface_grafica.services.export_service import ExportService


def test_export_excel_creates_file(tmp_path: Path) -> None:
    svc = ExportService()
    df = pl.DataFrame({"cnpj": ["12345678000190"], "valor": [100.0]})
    target = tmp_path / "relatorio.xlsx"
    result = svc.export_excel(target, df)
    assert result == target
    assert target.exists()


def test_export_excel_formats_dt_doc_and_dt_e_s_as_dates(tmp_path: Path) -> None:
    svc = ExportService()
    df = pl.DataFrame(
        {
            "Dt_doc": ["01/10/2021 00:00:00"],
            "Dt_e_s": ["01/10/2021 00:00:00"],
            "Cfop": ["1409"],
        }
    )
    target = tmp_path / "mov_estoque.xlsx"

    svc.export_excel(target, df, sheet_name="mov_estoque")

    ws = load_workbook(target).active
    assert ws["A2"].value.date() == date(2021, 10, 1)
    assert ws["B2"].value.date() == date(2021, 10, 1)
    assert ws["A2"].number_format == "dd/mm/yyyy"
    assert ws["B2"].number_format == "dd/mm/yyyy"


def test_export_txt_with_html(tmp_path: Path) -> None:
    svc = ExportService()
    target = tmp_path / "relatorio.html"
    html = "<html><body>Teste</body></html>"
    result = svc.export_txt_with_html(target, html)
    assert result == target
    assert target.exists()
    assert target.read_text(encoding="utf-8") == html


def test_export_docx_small_df(tmp_path: Path) -> None:
    svc = ExportService()
    df = pl.DataFrame({"cnpj": ["12345678000190"], "valor": [100.0]})
    target = tmp_path / "relatorio.docx"
    result = svc.export_docx(
        target=target,
        title="Relatório Teste",
        cnpj="12345678000190",
        table_name="tabela_teste",
        df=df,
        filters_text="",
        visible_columns=["cnpj", "valor"],
    )
    assert result == target
    assert target.exists()


def test_export_docx_large_df(tmp_path: Path) -> None:
    svc = ExportService()
    rows = 501
    df = pl.DataFrame({"id": list(range(rows)), "valor": [1.0] * rows})
    target = tmp_path / "relatorio_grande.docx"
    result = svc.export_docx(
        target=target,
        title="Relatório Grande",
        cnpj="12345678000190",
        table_name="tabela_grande",
        df=df,
        filters_text="periodo=2025",
        visible_columns=[],
    )
    assert result == target
    assert target.exists()
